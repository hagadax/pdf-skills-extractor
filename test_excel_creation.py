#!/usr/bin/env python3
"""
Script to create a test Excel file with skills data for testing the Excel extraction functionality.
"""

import openpyxl
from datetime import datetime
import os

def create_test_excel():
    """Create a test Excel file with technology skills."""
    
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Get the active worksheet
    ws = wb.active
    ws.title = "Skills Resume"
    
    # Add some sample data with technology skills
    data = [
        ["John Doe - Software Engineer Resume"],
        [""],
        ["Contact Information"],
        ["Email: john.doe@example.com"],
        ["Phone: (555) 123-4567"],
        [""],
        ["Technical Skills"],
        ["Programming Languages: Python, JavaScript, Java, C#, TypeScript"],
        ["Web Technologies: React, Vue.js, Angular, Node.js, Express.js"],
        ["Databases: PostgreSQL, MongoDB, MySQL, Redis"],
        ["Cloud Platforms: AWS, Azure, Google Cloud"],
        ["DevOps Tools: Docker, Kubernetes, Jenkins, Git, GitHub"],
        ["Frameworks: Django, Flask, Spring Boot, ASP.NET"],
        ["Testing: Jest, Pytest, Selenium, JUnit"],
        ["Other: RESTful APIs, GraphQL, Microservices, Agile, Scrum"],
        [""],
        ["Work Experience"],
        ["Senior Software Developer - Tech Corp (2020-Present)"],
        ["- Developed web applications using React and Node.js"],
        ["- Implemented CI/CD pipelines using Jenkins and Docker"],
        ["- Worked with microservices architecture on AWS"],
        ["- Used PostgreSQL and MongoDB for data storage"],
        [""],
        ["Software Engineer - StartupXYZ (2018-2020)"],
        ["- Built REST APIs using Python Django"],
        ["- Frontend development with Vue.js and TypeScript"],
        ["- Database design and optimization with MySQL"],
        ["- Implemented automated testing with Pytest and Jest"],
        [""],
        ["Education"],
        ["Bachelor of Science in Computer Science"],
        ["University of Technology (2014-2018)"],
        [""],
        ["Certifications"],
        ["AWS Certified Solutions Architect"],
        ["Azure Fundamentals"],
        ["Kubernetes Administrator"],
    ]
    
    # Add data to worksheet
    for row_idx, row_data in enumerate(data, 1):
        ws[f'A{row_idx}'] = row_data[0]
    
    # Create a second sheet with project details
    ws2 = wb.create_sheet("Projects")
    project_data = [
        ["Project Name", "Technologies Used", "Description"],
        ["E-commerce Platform", "React, Node.js, MongoDB, Docker", "Full-stack web application"],
        ["Data Analytics Tool", "Python, Pandas, Flask, PostgreSQL", "Data processing and visualization"],
        ["Mobile App Backend", "Java, Spring Boot, MySQL, Redis", "RESTful API for mobile application"],
        ["ML Pipeline", "Python, TensorFlow, Kubernetes, AWS", "Machine learning model deployment"],
        ["Chat Application", "Vue.js, Socket.io, Express.js, MongoDB", "Real-time messaging system"]
    ]
    
    for row_idx, row_data in enumerate(project_data, 1):
        for col_idx, cell_data in enumerate(row_data, 1):
            ws2.cell(row=row_idx, column=col_idx, value=cell_data)
    
    # Set document properties
    wb.properties.title = "Software Engineer Resume"
    wb.properties.subject = "Resume with Technical Skills"
    wb.properties.creator = "Skills Extractor Test"
    wb.properties.created = datetime.now()
    
    # Save the file
    filename = "test_resume_with_skills.xlsx"
    filepath = os.path.join("/Users/a.abdurihim/Documents/Code/get-skills", filename)
    wb.save(filepath)
    
    print(f"Test Excel file created: {filepath}")
    return filepath

if __name__ == "__main__":
    create_test_excel()
