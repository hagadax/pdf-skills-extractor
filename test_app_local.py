#!/usr/bin/env python3
"""
Test local version of the GET-SKILLS application without Azure dependencies
"""

import os
import json
from flask import Flask, request, render_template, jsonify, redirect, url_for
from werkzeug.utils import secure_filename
from collections import Counter, defaultdict
from datetime import datetime
import PyPDF2
import openpyxl
import xlrd

# Import AI skills module
from ai_skills import AISkillExtractor

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Global variables for local testing
skill_counter = Counter()
skill_documents = defaultdict(list)
processed_documents = []
monthly_skill_data = defaultdict(lambda: defaultdict(int))

# AI extraction variables
ai_skill_counter = Counter()
ai_skill_documents = defaultdict(list)
ai_processed_documents = []

# Initialize AI extractor
ai_extractor = AISkillExtractor()

# Skills list (subset for testing)
SKILLS = [
    'Python', 'JavaScript', 'Java', 'C++', 'React', 'Node.js', 'Docker', 'Kubernetes',
    'AWS', 'Azure', 'Git', 'SQL', 'MongoDB', 'Flask', 'Django', 'Express',
    'HTML', 'CSS', 'TypeScript', 'Angular', 'Vue.js', 'PostgreSQL', 'MySQL',
    'Redis', 'Jenkins', 'CI/CD', 'DevOps', 'Linux', 'Agile', 'Scrum'
]

def extract_text_from_pdf(file_path):
    """Extract text from PDF file"""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""

def extract_text_from_excel(file_path):
    """Extract text from Excel file"""
    try:
        if file_path.endswith('.xlsx'):
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text += str(cell) + " "
            return text
        else:
            workbook = xlrd.open_workbook(file_path)
            text = ""
            for sheet in workbook.sheets():
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        cell_value = sheet.cell_value(row, col)
                        if cell_value:
                            text += str(cell_value) + " "
            return text
    except Exception as e:
        print(f"Error extracting text from Excel: {e}")
        return ""

def extract_skills_from_text(text, filename):
    """Extract skills using pattern matching"""
    found_skills = []
    text_lower = text.lower()
    
    for skill in SKILLS:
        if skill.lower() in text_lower:
            found_skills.append(skill)
            skill_counter[skill] += 1
            skill_documents[skill].append(filename)
    
    return found_skills

@app.route('/')
def index():
    # Prepare chart data for the template
    chart_data = {
        'datasets': [],
        'labels': []
    }
    
    # Generate last 12 months labels
    from datetime import datetime, timedelta
    import calendar
    
    now = datetime.now()
    months = []
    for i in range(11, -1, -1):
        month_date = now - timedelta(days=30*i)
        months.append(calendar.month_abbr[month_date.month] + ' ' + str(month_date.year))
    
    chart_data['labels'] = months
    
    # If we have skills data, create datasets for top skills
    if skill_counter:
        top_skills = skill_counter.most_common(5)  # Top 5 skills for chart
        colors = ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF']
        
        for i, (skill, _) in enumerate(top_skills):
            # Generate dummy monthly data for demo
            import random
            monthly_data = [random.randint(0, 10) for _ in range(12)]
            
            chart_data['datasets'].append({
                'label': skill,
                'data': monthly_data,
                'borderColor': colors[i % len(colors)],
                'backgroundColor': colors[i % len(colors)] + '20',
                'fill': False,
                'tension': 0.4
            })
    
    return render_template('index.html', 
                         page_name='home', 
                         chart_data=chart_data,
                         total_skills=len(skill_counter),
                         total_documents=len(processed_documents),
                         recent_documents=processed_documents[-5:] if processed_documents else [])

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file selected'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'})
    
    if file:
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        
        # Extract text based on file type
        if filename.lower().endswith('.pdf'):
            text = extract_text_from_pdf(file_path)
        elif filename.lower().endswith(('.xlsx', '.xls')):
            text = extract_text_from_excel(file_path)
        else:
            return jsonify({'error': 'Unsupported file type'})
        
        if not text:
            return jsonify({'error': 'Could not extract text from file'})
        
        # Pattern matching extraction
        pattern_skills = extract_skills_from_text(text, filename)
        
        # AI extraction (if API key available)
        ai_skills = []
        try:
            # Determine document type from filename
            doc_type = "resume" if any(term in filename.lower() for term in ["cv", "resume"]) else "job_description"
            ai_skills, ai_metadata = ai_extractor.extract_skills_from_text(text, doc_type)
            print(f"AI extracted skills: {ai_skills}")
        except Exception as e:
            print(f"AI extraction failed: {e}")
        
        # Update processed documents
        processed_documents.append({
            'filename': filename,
            'date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'pattern_skills_count': len(pattern_skills),
            'ai_skills_count': len(ai_skills)
        })
        
        return jsonify({
            'success': True,
            'filename': filename,
            'pattern_skills': pattern_skills,
            'ai_skills': ai_skills,
            'pattern_skills_count': len(pattern_skills),
            'ai_skills_count': len(ai_skills)
        })

@app.route('/skills')
def skills():
    skills_list = [(skill, count) for skill, count in skill_counter.most_common()]
    return render_template('skills.html', skills=skills_list, page_name='skills')

@app.route('/ai-skills')
def ai_skills():
    ai_skills_list = ai_extractor.get_ai_skills_stats()
    return render_template('ai_skills.html', ai_skills=ai_skills_list, page_name='ai-skills')

@app.route('/comparison')
def comparison():
    # Get both skill sets
    pattern_skills = [(skill, count) for skill, count in skill_counter.most_common()]
    ai_skills_list = ai_extractor.get_ai_skills_stats()
    
    # Calculate comparison statistics
    pattern_skill_names = set([skill for skill, _ in pattern_skills])
    ai_skill_names = set([skill for skill, _ in ai_skills_list])
    
    overlap_count = len(pattern_skill_names.intersection(ai_skill_names))
    total_unique = len(pattern_skill_names.union(ai_skill_names))
    unique_pattern_skills = list(pattern_skill_names - ai_skill_names)
    unique_ai_skills = list(ai_skill_names - pattern_skill_names)
    
    return render_template('comparison.html', 
                         pattern_skills=pattern_skills,
                         ai_skills=ai_skills_list,
                         overlap_count=overlap_count,
                         total_unique=total_unique,
                         unique_pattern_skills=unique_pattern_skills,
                         unique_ai_skills=unique_ai_skills,
                         page_name='comparison')

@app.route('/documents')
def documents():
    return render_template('documents.html', documents=processed_documents, page_name='documents')

@app.route('/about')
def about():
    return render_template('about.html', page_name='about')

@app.route('/api/stats')
def api_stats():
    return jsonify({
        'total_skills': len(skill_counter),
        'total_documents': len(processed_documents),
        'top_skills': dict(skill_counter.most_common(10))
    })

@app.route('/api/ai-stats')
def api_ai_stats():
    ai_stats = ai_extractor.get_ai_skills_stats()
    return jsonify({
        'total_ai_skills': len(ai_stats),
        'ai_skills': dict(ai_stats)
    })

if __name__ == '__main__':
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    print("Starting GET-SKILLS application in local test mode...")
    
    # Check AI configuration
    if hasattr(ai_extractor, 'service_type'):
        print(f"✅ AI Service: {ai_extractor.service_type}")
        if ai_extractor.service_type == "Azure OpenAI":
            print(f"   Model: {ai_extractor.model_name}")
            print(f"   Endpoint: {os.environ.get('AZURE_OPENAI_ENDPOINT', 'Not set')}")
        else:
            print(f"   Model: {ai_extractor.model_name}")
    else:
        print("⚠️  No AI service configured")
        print("   Set either:")
        print("   - Azure OpenAI: AZURE_OPENAI_ENDPOINT, AZURE_OPENAI_API_KEY, AZURE_OPENAI_DEPLOYMENT_NAME")
        print("   - OpenAI: OPENAI_API_KEY")
        print("   See AZURE_AI_SETUP.md for detailed setup instructions")
    
    app.run(debug=True, host='0.0.0.0', port=5000)