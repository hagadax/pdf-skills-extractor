from flask import Flask, request, render_template, redirect, url_for, flash, jsonify
import os
import re
import PyPDF2
import openpyxl
import xlrd
import pandas as pd
from werkzeug.utils import secure_filename
from collections import defaultdict, Counter
from datetime import datetime
import io
import tempfile
from azure.storage.blob import BlobServiceClient
from azure.identity import DefaultAzureCredential
from skills import extract_skills, skill_counter, monthly_skill_data, skill_documents, processed_documents, tech_skills
from ai_skills import ai_extractor
from monthly_analysis import monthly_analyzer
from keyvault_manager import get_application_config
import json
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

# Get configuration from Key Vault with environment fallbacks
config = get_application_config()

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here')
app.config['UPLOAD_FOLDER'] = os.environ.get('UPLOAD_FOLDER', 'uploads')
app.config['MAX_CONTENT_LENGTH'] = int(os.environ.get('MAX_CONTENT_LENGTH', 16 * 1024 * 1024))

# Azure Blob Storage configuration from Key Vault
AZURE_STORAGE_CONNECTION_STRING = config.get('azure_storage_connection_string')
AZURE_STORAGE_CONTAINER_NAME = os.environ.get('AZURE_STORAGE_CONTAINER_NAME', 'uploads')

# Initialize Azure Blob Service Client with Managed Identity
def get_blob_service_client():
    """Initialize Azure Blob Service Client using Managed Identity or connection string."""
    try:
        if AZURE_STORAGE_CONNECTION_STRING:
            return BlobServiceClient.from_connection_string(AZURE_STORAGE_CONNECTION_STRING)
        else:
            # Use Managed Identity for authentication
            credential = DefaultAzureCredential()
            account_url = f"https://{os.environ.get('AZURE_STORAGE_ACCOUNT_NAME')}.blob.core.windows.net"
            return BlobServiceClient(account_url=account_url, credential=credential)
    except Exception as e:
        print(f"Error initializing blob service client: {e}")
        return None

# Ensure upload directory exists (fallback for local development)
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global skill tracker
skill_counter = Counter()

# Track skills by document with metadata
# Structure: {skill: [{filename, upload_date, file_date}, ...]}
skill_documents = defaultdict(list)

# Track all processed documents
# Structure: {filename: {upload_date, file_date, skills_found}}
processed_documents = {}

# Track monthly skill data for charts
# Structure: {skill: {'2025-01': 5, '2025-02': 3, ...}}
monthly_skill_data = defaultdict(lambda: defaultdict(int))

# Stats persistence configuration
STATS_BLOB_NAME = 'app_stats.json'
STATS_CONTAINER_NAME = os.environ.get('AZURE_STORAGE_CONTAINER_NAME', 'uploads')

def save_stats_to_blob():
    """Save application statistics to Azure Blob Storage as JSON."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            print("Warning: Azure Blob Storage not available for stats persistence")
            return False
        
        # Prepare stats data for serialization
        stats_data = {
            'skill_counter': dict(skill_counter),
            'skill_documents': dict(skill_documents),
            'processed_documents': processed_documents,
            'monthly_skill_data': {
                skill: dict(months) for skill, months in monthly_skill_data.items()
            },
            'last_updated': datetime.now().isoformat(),
            'version': '1.0'  # For future compatibility
        }
        
        # Convert to JSON
        stats_json = json.dumps(stats_data, indent=2, default=str)
        
        # Upload to blob storage
        blob_client = blob_service_client.get_blob_client(
            container=STATS_CONTAINER_NAME,
            blob=STATS_BLOB_NAME
        )
        
        blob_client.upload_blob(stats_json, overwrite=True)
        print(f"Stats saved to blob storage successfully at {datetime.now()}")
        
        # Also save AI statistics
        save_ai_stats_to_blob()
        
        return True
        
    except Exception as e:
        print(f"Error saving stats to blob storage: {e}")
        return False

def save_ai_stats_to_blob():
    """Save AI extraction statistics to Azure Blob Storage."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            print("Warning: No blob service client available for AI stats save")
            return False
        
        ai_stats_data = ai_extractor.get_ai_stats_data()
        print(f"Saving AI stats: {len(ai_stats_data.get('ai_skill_counter', {}))} skills, {len(ai_stats_data.get('ai_processed_documents', {}))} documents")
        ai_stats_json = json.dumps(ai_stats_data, indent=2, default=str)
        
        # Upload AI stats to blob storage
        ai_blob_client = blob_service_client.get_blob_client(
            container=STATS_CONTAINER_NAME,
            blob=ai_extractor.ai_stats_blob_name
        )
        
        ai_blob_client.upload_blob(ai_stats_json, overwrite=True)
        print(f"AI stats saved to blob storage successfully at {datetime.now()}")
        return True
        
    except Exception as e:
        print(f"Error saving AI stats to blob storage: {e}")
        import traceback
        traceback.print_exc()
        return False

def load_stats_from_blob():
    """Load application statistics from Azure Blob Storage."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            print("Info: Azure Blob Storage not available, starting with empty stats")
            return False
        
        blob_client = blob_service_client.get_blob_client(
            container=STATS_CONTAINER_NAME,
            blob=STATS_BLOB_NAME
        )
        
        # Check if stats file exists
        if not blob_client.exists():
            print("Info: No existing stats found, starting with empty stats")
            return False
        
        # Download and parse stats
        stats_content = blob_client.download_blob().readall()
        stats_data = json.loads(stats_content.decode('utf-8'))
        
        # Restore global variables
        global skill_counter, skill_documents, processed_documents, monthly_skill_data
        
        skill_counter = Counter(stats_data.get('skill_counter', {}))
        
        # Restore skill_documents (convert back to defaultdict)
        skill_documents_data = stats_data.get('skill_documents', {})
        skill_documents = defaultdict(list)
        for skill, docs in skill_documents_data.items():
            skill_documents[skill] = docs
        
        processed_documents = stats_data.get('processed_documents', {})
        
        # Restore monthly_skill_data (convert back to nested defaultdict)
        monthly_data = stats_data.get('monthly_skill_data', {})
        monthly_skill_data = defaultdict(lambda: defaultdict(int))
        for skill, months in monthly_data.items():
            monthly_skill_data[skill] = defaultdict(int, months)
        
        last_updated = stats_data.get('last_updated', 'Unknown')
        version = stats_data.get('version', 'Unknown')
        
        print(f"Stats loaded from blob storage successfully")
        print(f"  Last updated: {last_updated}")
        print(f"  Version: {version}")
        print(f"  Documents: {len(processed_documents)}")
        print(f"  Skills: {len(skill_counter)}")
        
        # Also load AI statistics
        load_ai_stats_from_blob()
        
        return True
        
    except Exception as e:
        print(f"Error loading stats from blob storage: {e}")
        import traceback
        traceback.print_exc()
        return False

def sync_ai_extractor_with_processed_documents():
    """Sync AI extractor counters with data from processed documents."""
    try:
        print("Syncing AI extractor with processed documents...")
        
        # Only clear and rebuild if we have no AI data or if explicitly requested
        needs_rebuild = (
            len(ai_extractor.ai_skill_counter) == 0 or 
            len(ai_extractor.ai_processed_documents) == 0
        )
        
        if not needs_rebuild:
            print("AI extractor already has data, skipping sync to preserve existing state")
            return True
            
        ai_extractor.ai_skill_counter.clear()
        ai_extractor.ai_processed_documents.clear()
        ai_extractor.ai_skill_documents.clear()
        ai_extractor.ai_monthly_skill_data.clear()
        
        # Rebuild AI stats from processed documents
        rebuilt_count = 0
        for filename, doc_data in processed_documents.items():
            ai_skills = doc_data.get('ai_skills_found', [])
            if ai_skills:
                rebuilt_count += 1
                # Update AI skill counter
                for skill in ai_skills:
                    ai_extractor.ai_skill_counter[skill] += 1
                    ai_extractor.ai_skill_documents[skill].append(filename)
                
                # Update processed documents
                ai_extractor.ai_processed_documents[filename] = {
                    'skills': ai_skills,
                    'processed_at': doc_data.get('upload_date', datetime.now().isoformat()),
                    'skill_count': len(ai_skills),
                    'file_type': doc_data.get('file_type', 'unknown')
                }
                
                # Update monthly data 
                file_date = doc_data.get('file_date') or doc_data.get('upload_date', '').split(' ')[0]
                if file_date:
                    month_key = file_date[:7]  # YYYY-MM format
                    for skill in ai_skills:
                        ai_extractor.ai_monthly_skill_data[skill][month_key] += 1
        
        print(f"AI extractor sync complete: {len(ai_extractor.ai_skill_counter)} skills, {len(ai_extractor.ai_processed_documents)} documents from {rebuilt_count} files with AI data")
        
        # Save the rebuilt AI stats to ensure persistence
        if len(ai_extractor.ai_skill_counter) > 0:
            save_ai_stats_to_blob()
            
        return True
        
    except Exception as e:
        print(f"Error syncing AI extractor: {e}")
        import traceback
        traceback.print_exc()
        return False

def force_sync_ai_extractor_with_processed_documents():
    """Force sync AI extractor counters with data from processed documents."""
    try:
        print("Force syncing AI extractor with processed documents...")
        ai_extractor.ai_skill_counter.clear()
        ai_extractor.ai_processed_documents.clear()
        ai_extractor.ai_skill_documents.clear()
        ai_extractor.ai_monthly_skill_data.clear()
        
        # Rebuild AI stats from processed documents
        rebuilt_count = 0
        for filename, doc_data in processed_documents.items():
            ai_skills = doc_data.get('ai_skills_found', [])
            if ai_skills:
                rebuilt_count += 1
                # Update AI skill counter
                for skill in ai_skills:
                    ai_extractor.ai_skill_counter[skill] += 1
                    ai_extractor.ai_skill_documents[skill].append(filename)
                
                # Update processed documents
                ai_extractor.ai_processed_documents[filename] = {
                    'skills': ai_skills,
                    'processed_at': doc_data.get('upload_date', datetime.now().isoformat()),
                    'skill_count': len(ai_skills),
                    'file_type': doc_data.get('file_type', 'unknown')
                }
                
                # Update monthly data 
                file_date = doc_data.get('file_date') or doc_data.get('upload_date', '').split(' ')[0]
                if file_date:
                    month_key = file_date[:7]  # YYYY-MM format
                    for skill in ai_skills:
                        ai_extractor.ai_monthly_skill_data[skill][month_key] += 1
        
        print(f"Force AI extractor sync complete: {len(ai_extractor.ai_skill_counter)} skills, {len(ai_extractor.ai_processed_documents)} documents from {rebuilt_count} files with AI data")
        
        # Save the rebuilt AI stats to ensure persistence
        if len(ai_extractor.ai_skill_counter) > 0:
            save_ai_stats_to_blob()
            
        return True
        
    except Exception as e:
        print(f"Error force syncing AI extractor: {e}")
        import traceback
        traceback.print_exc()
        return False

def load_ai_stats_from_blob():
    """Load AI extraction statistics from Azure Blob Storage."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            print("Warning: No blob service client available for AI stats load")
            return False
        
        ai_blob_client = blob_service_client.get_blob_client(
            container=STATS_CONTAINER_NAME,
            blob=ai_extractor.ai_stats_blob_name
        )
        
        if not ai_blob_client.exists():
            print("Info: No existing AI stats found, starting with empty AI stats")
            return False
        
        ai_stats_content = ai_blob_client.download_blob().readall()
        ai_stats_data = json.loads(ai_stats_content.decode('utf-8'))
        
        print(f"Loading AI stats from blob: {len(ai_stats_data.get('ai_skill_counter', {}))} skills, {len(ai_stats_data.get('ai_processed_documents', {}))} documents")
        
        ai_extractor.load_ai_stats_data(ai_stats_data)
        
        print(f"AI stats loaded from blob storage successfully")
        print(f"  AI Documents: {len(ai_extractor.ai_processed_documents)}")
        print(f"  AI Skills: {len(ai_extractor.ai_skill_counter)}")
        print(f"  Top AI Skills: {ai_extractor.ai_skill_counter.most_common(5)}")
        
        return True
        
    except Exception as e:
        print(f"Error loading AI stats from blob storage: {e}")
        import traceback
        traceback.print_exc()
        return False

# Load stats on application startup
try:
    stats_loaded = load_stats_from_blob()
    if stats_loaded:
        print("Application stats loaded successfully on startup")
        # Sync AI extractor with processed documents data
        sync_ai_extractor_with_processed_documents()
        # Also try to load dedicated AI stats to fill any gaps
        if len(ai_extractor.ai_skill_counter) == 0:
            print("AI stats still empty after sync, attempting explicit AI stats load...")
            load_ai_stats_from_blob()
    else:
        print("Failed to load main stats, attempting AI stats load only...")
        load_ai_stats_from_blob()
except Exception as e:
    print(f"Failed to load stats on startup: {e}")
    print("Starting with empty stats")
    # Still try to load AI stats independently
    try:
        load_ai_stats_from_blob()
    except Exception as ai_e:
        print(f"Also failed to load AI stats: {ai_e}")

# Allowed file extensions
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls'}

def allowed_file(filename):
    """Check if the uploaded file has an allowed extension."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_text_from_pdf(file_content):
    """Extract text content from a PDF file (from bytes or file path)."""
    try:
        if isinstance(file_content, str):
            # Legacy support for file paths
            with open(file_content, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
        else:
            # Handle bytes content from blob storage
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        
        text = ""
        for page in pdf_reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        
        # Clean up common PDF extraction issues
        # Remove excessive whitespace while preserving word boundaries
        text = re.sub(r'\s+', ' ', text)  # Replace multiple spaces/newlines with single space
        text = re.sub(r'([a-z])([A-Z])', r'\1 \2', text)  # Add space between camelCase
        text = text.strip()
        
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""

def get_pdf_creation_date(file_content):
    """Extract creation date from PDF metadata (from bytes or file path)."""
    try:
        if isinstance(file_content, str):
            # Legacy support for file paths
            with open(file_content, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
        else:
            # Handle bytes content from blob storage
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
        
        if pdf_reader.metadata and '/CreationDate' in pdf_reader.metadata:
            creation_date = pdf_reader.metadata.get('/CreationDate')
            if creation_date:
                # PDF dates are in format D:YYYYMMDDHHmmSSOHH'mm
                date_str = str(creation_date)
                if date_str.startswith('D:'):
                    date_str = date_str[2:16]  # Extract YYYYMMDDHHMMSS
                    try:
                        return datetime.strptime(date_str[:8], '%Y%m%d').strftime('%Y-%m-%d')
                    except:
                        pass
        return None
    except Exception as e:
        print(f"Error extracting PDF date: {e}")
        return None

def extract_text_from_excel(file_content, filename):
    """Extract text content from an Excel file (from bytes)."""
    try:
        # Determine file type by extension
        file_extension = filename.lower().split('.')[-1]
        
        # Create a BytesIO object from the file content
        excel_file = io.BytesIO(file_content)
        
        # Read Excel file based on extension
        if file_extension == 'xlsx':
            # Use openpyxl for .xlsx files
            workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            text_content = []
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text_content.append(str(cell))
            
            workbook.close()
            
        elif file_extension == 'xls':
            # Use xlrd for .xls files
            workbook = xlrd.open_workbook(file_contents=file_content)
            text_content = []
            
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                for row_index in range(sheet.nrows):
                    for col_index in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_index, col_index)
                        if cell_value:
                            text_content.append(str(cell_value))
        else:
            return ""
        
        # Join all text with spaces and clean up
        text = " ".join(text_content)
        
        # Clean up common Excel extraction issues
        text = re.sub(r'\s+', ' ', text)  # Replace multiple spaces with single space
        text = text.strip()
        
        return text
        
    except Exception as e:
        print(f"Error extracting text from Excel file: {e}")
        return ""

def get_excel_creation_date(file_content, filename):
    """Extract creation date from Excel metadata (from bytes)."""
    try:
        file_extension = filename.lower().split('.')[-1]
        excel_file = io.BytesIO(file_content)
        
        if file_extension == 'xlsx':
            # Use openpyxl for .xlsx files
            workbook = openpyxl.load_workbook(excel_file, read_only=True)
            
            # Try to get creation date from document properties
            if hasattr(workbook, 'properties') and workbook.properties:
                if hasattr(workbook.properties, 'created') and workbook.properties.created:
                    return workbook.properties.created.strftime('%Y-%m-%d')
                elif hasattr(workbook.properties, 'modified') and workbook.properties.modified:
                    return workbook.properties.modified.strftime('%Y-%m-%d')
            
            workbook.close()
            
        elif file_extension == 'xls':
            # For .xls files, xlrd doesn't provide easy access to metadata
            # We could use pandas as an alternative but it's less reliable for dates
            pass
            
        return None
        
    except Exception as e:
        print(f"Error extracting Excel date: {e}")
        return None

def get_file_type(filename):
    """Determine file type based on extension."""
    extension = filename.lower().split('.')[-1]
    if extension == 'pdf':
        return 'pdf'
    elif extension in ['xlsx', 'xls']:
        return 'excel'
    else:
        return 'unknown'

def extract_skills(text):
    """Extract technology skills from text using improved pattern matching.
    Handles PDF text fragmentation and skill variations."""
    found_skills = set()  # Use set to ensure uniqueness
    
    # Normalize text: remove excessive whitespace and line breaks
    normalized_text = ' '.join(text.split()).lower()
    
    # Also create a version without spaces for compound skills
    no_spaces_text = re.sub(r'\s+', '', text.lower())
    
    for skill in tech_skills:
        skill_lower = skill.lower()
        skill_found = False
        
        # Method 1: Standard word boundary matching (for simple skills)
        pattern = r'\b' + re.escape(skill_lower) + r'\b'
        if re.search(pattern, normalized_text):
            skill_found = True
        
        # Method 2: Flexible matching for compound skills (allows spaces/dots)
        # Handle skills like "Node.js", "Vue.js", "C++", "C#", etc.
        flexible_pattern = re.escape(skill_lower)
        flexible_pattern = flexible_pattern.replace(r'\.', r'[\.\s]*')  # . or spaces
        flexible_pattern = flexible_pattern.replace(r'\+', r'[\+\s]*')  # + or spaces  
        flexible_pattern = flexible_pattern.replace(r'\#', r'[\#\s]*')  # # or spaces
        flexible_pattern = flexible_pattern.replace(r'\s', r'\s*')      # flexible spaces
        
        if re.search(r'\b' + flexible_pattern + r'\b', normalized_text):
            skill_found = True
        
        # Method 3: No-spaces matching for fragmented text
        skill_no_spaces = re.sub(r'[\s\.\+\#-]', '', skill_lower)
        if len(skill_no_spaces) > 2 and skill_no_spaces in no_spaces_text:
            skill_found = True
        
        if skill_found:
            found_skills.add(skill)
    
    # Additional skill variations and aliases
    skill_aliases = {
        'js': 'JavaScript',
        'ts': 'TypeScript', 
        'nodejs': 'Node.js',
        'reactjs': 'React',
        'vuejs': 'Vue.js',
        'nextjs': 'Next.js',
        'nuxtjs': 'Nuxt.js',
        'dotnet': 'ASP.NET',
        '.net': 'ASP.NET',
        'csharp': 'C#',
        'cplusplus': 'C++',
        'ai': 'Artificial Intelligence',
        'ml': 'Machine Learning',
        'dl': 'Deep Learning',
        'k8s': 'Kubernetes',
        'aws': 'AWS',
        'gcp': 'Google Cloud',
        'azure': 'Azure',
        'vscode': 'Visual Studio Code',
        'vs code': 'Visual Studio Code',
        'sql server': 'SQL Server',
        'postgresql': 'PostgreSQL',
        'mongo': 'MongoDB',
        'redis': 'Redis',
        'docker': 'Docker',
        'git': 'Git',
        'github': 'GitHub',
        'gitlab': 'GitLab',
        'rest': 'REST API',
        'api': 'API Development',
        'ci/cd': 'CI/CD',
        'cicd': 'CI/CD'
    }
    
    # Check for aliases in normalized text
    for alias, skill_name in skill_aliases.items():
        # Only add if the main skill wasn't already found
        if skill_name not in found_skills:
            alias_pattern = r'\b' + re.escape(alias) + r'\b'
            if re.search(alias_pattern, normalized_text):
                found_skills.add(skill_name)
    
    return list(found_skills)  # Convert back to list for consistency

def get_monthly_chart_data():
    """Generate chart data for top 10 skills over the past 12 months."""
    from datetime import datetime, timedelta
    import calendar
    
    # Get current date and calculate 12 months back
    current_date = datetime.now()
    months = []
    
    # Generate list of last 12 months in YYYY-MM format
    for i in range(11, -1, -1):
        date = current_date - timedelta(days=i*30)  # Approximate month calculation
        month_key = date.strftime('%Y-%m')
        month_label = date.strftime('%b %Y')
        months.append({'key': month_key, 'label': month_label})
    
    # Get top 10 skills
    top_skills = [skill for skill, count in skill_counter.most_common(10)]
    
    # Prepare chart data
    chart_data = {
        'labels': [month['label'] for month in months],
        'datasets': []
    }
    
    # Color palette for the chart
    colors = [
        '#667eea', '#764ba2', '#f093fb', '#f5576c', '#4facfe', 
        '#43e97b', '#fa709a', '#fee140', '#a8edea', '#d299c2'
    ]
    
    # Create dataset for each top skill
    for i, skill in enumerate(top_skills):
        skill_data = []
        cumulative_count = 0
        
        for month in months:
            # Add monthly occurrences to cumulative count
            monthly_occurrences = monthly_skill_data[skill].get(month['key'], 0)
            cumulative_count += monthly_occurrences
            skill_data.append(cumulative_count)
        
        # Only include skills that have data
        if cumulative_count > 0:
            chart_data['datasets'].append({
                'label': skill,
                'data': skill_data,
                'borderColor': colors[i % len(colors)],
                'backgroundColor': colors[i % len(colors)] + '20',  # 20 for transparency
                'tension': 0.4,
                'fill': False
            })
    
    return chart_data

@app.route('/')
def index():
    """Main page with upload form and skill statistics."""
    # Get top 10 most common skills from pattern matching
    top_skills = skill_counter.most_common(10)
    
    # Get top 10 most common AI skills
    top_ai_skills = ai_extractor.ai_skill_counter.most_common(10)
    
    total_documents = len(processed_documents)
    total_ai_documents = len(ai_extractor.ai_processed_documents)
    
    # Get chart data for visualization (pattern matching)
    pattern_chart_data = get_monthly_chart_data()
    
    # Get AI chart data for visualization
    ai_chart_data = ai_extractor.get_ai_monthly_chart_data()
    
    return render_template('index.html', 
                         top_skills=top_skills, 
                         top_ai_skills=top_ai_skills,
                         total_documents=total_documents,
                         total_ai_documents=total_ai_documents,
                         total_skills=len(skill_counter),
                         total_ai_skills=len(ai_extractor.ai_skill_counter),
                         pattern_chart_data=pattern_chart_data,
                         ai_chart_data=ai_chart_data,
                         page_name='home')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle PDF and Excel file upload and skill extraction."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file selected'})
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        
        try:
            # Save file to Azure Blob Storage or local storage
            file_content, is_blob = save_file_safely(file, filename)
            
            # Get upload date
            upload_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Determine file type and extract content accordingly
            file_type = get_file_type(filename)
            
            if file_type == 'pdf':
                # Extract text and metadata from PDF
                text = extract_text_from_pdf(file_content)
                file_date = get_pdf_creation_date(file_content)
            elif file_type == 'excel':
                # Extract text and metadata from Excel
                text = extract_text_from_excel(file_content, filename)
                file_date = get_excel_creation_date(file_content, filename)
            else:
                return jsonify({'success': False, 'message': 'Unsupported file type'})
            
            if text:
                # Extract skills from text using pattern matching (existing method)
                found_skills = extract_skills(text)
                
                # Extract skills using AI (new method)
                ai_skills = []
                ai_metadata = {}
                try:
                    # Determine document type from filename
                    doc_type = "resume" if any(term in filename.lower() for term in ["cv", "resume"]) else "job_description"
                    ai_skills, ai_metadata = ai_extractor.extract_skills_from_text(text, doc_type)
                    print(f"AI extracted {len(ai_skills)} skills: {ai_skills}")
                    
                    # Update AI extractor's internal state immediately
                    if ai_skills:
                        for skill in ai_skills:
                            ai_extractor.ai_skill_counter[skill] += 1
                            ai_extractor.ai_skill_documents[skill].append(filename)
                            # Update monthly data 
                            date_for_tracking = file_date if file_date else upload_date.split(' ')[0]
                            month_key = date_for_tracking[:7]  # Get YYYY-MM format
                            ai_extractor.ai_monthly_skill_data[skill][month_key] += 1
                        
                        # Update AI processed documents
                        ai_extractor.ai_processed_documents[filename] = {
                            'skills': ai_skills,
                            'processed_at': upload_date,
                            'skill_count': len(ai_skills),
                            'file_type': file_type
                        }
                        
                except Exception as e:
                    print(f"AI extraction failed: {e}")
                    ai_metadata = {'error': str(e)}
                
                # Get month key for tracking (use file date if available, otherwise upload date)
                date_for_tracking = file_date if file_date else upload_date.split(' ')[0]
                month_key = date_for_tracking[:7]  # Get YYYY-MM format
                
                # Update global skill counter (pattern matching)
                for skill in found_skills:
                    skill_counter[skill] += 1  # +1 per document, regardless of skill frequency in text
                    
                    # Track monthly data
                    monthly_skill_data[skill][month_key] += 1
                    
                    # Track which document contains this skill
                    skill_documents[skill].append({
                        'filename': filename,
                        'upload_date': upload_date,
                        'file_date': file_date or upload_date.split(' ')[0],  # Use upload date if no file date
                        'file_type': file_type
                    })
                
                # Track processed document
                processed_documents[filename] = {
                    'upload_date': upload_date,
                    'file_date': file_date or upload_date.split(' ')[0],
                    'skills_found': found_skills,
                    'ai_skills_found': ai_skills,
                    'ai_metadata': ai_metadata,
                    'storage_type': 'blob' if is_blob else 'local',
                    'file_type': file_type
                }
                
                # Save stats to blob storage after processing
                save_stats_to_blob()
                
                # Also save a backup with timestamp for data recovery
                try:
                    backup_blob_name = f'backups/stats_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
                    blob_service_client = get_blob_service_client()
                    if blob_service_client:
                        backup_stats = {
                            'skill_counter': dict(skill_counter),
                            'processed_documents': processed_documents,
                            'backup_timestamp': datetime.now().isoformat()
                        }
                        backup_blob_client = blob_service_client.get_blob_client(
                            container=STATS_CONTAINER_NAME,
                            blob=backup_blob_name
                        )
                        backup_blob_client.upload_blob(
                            json.dumps(backup_stats, indent=2).encode('utf-8'), 
                            overwrite=True
                        )
                        print(f"Backup created: {backup_blob_name}")
                except Exception as backup_error:
                    print(f"Warning: Failed to create backup: {backup_error}")
                
                storage_msg = 'Azure Blob Storage' if is_blob else 'local storage'
                file_type_msg = 'PDF' if file_type == 'pdf' else 'Excel'
                
                # Create response message
                response_message = f'Successfully processed {file_type_msg} file {filename} (saved to {storage_msg}).\n'
                response_message += f'Pattern Matching found {len(found_skills)} skills: {", ".join(found_skills)}'
                
                if ai_skills:
                    response_message += f'\nAI Extraction found {len(ai_skills)} skills: {", ".join(ai_skills)}'
                elif 'error' in ai_metadata:
                    response_message += f'\nAI Extraction failed: {ai_metadata["error"]}'
                else:
                    response_message += f'\nAI Extraction: No OpenAI API key configured'
                
                return jsonify({
                    'success': True, 
                    'message': response_message,
                    'skills': found_skills,
                    'ai_skills': ai_skills,
                    'filename': filename,
                    'file_type': file_type,
                    'extraction_methods': {
                        'pattern_matching': len(found_skills),
                        'ai_extraction': len(ai_skills)
                    }
                })
            else:
                file_type_msg = 'PDF' if file_type == 'pdf' else 'Excel'
                return jsonify({'success': False, 'message': f'Could not extract text from {file_type_msg} file {filename}'})
                
        except Exception as e:
            return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'})
    else:
        return jsonify({'success': False, 'message': 'Please upload a PDF or Excel file (.pdf, .xlsx, .xls)'})

@app.route('/api/skills')
def api_skills():
    """API endpoint to get skill statistics as JSON."""
    return jsonify({
        'total_skills': len(skill_counter),
        'skills': dict(skill_counter),
        'top_skills': skill_counter.most_common(20)
    })

@app.route('/api/ai-skills')
def api_ai_skills():
    """API endpoint to get AI-extracted skill statistics as JSON."""
    return jsonify({
        'total_ai_skills': len(ai_extractor.ai_skill_counter),
        'ai_skills': dict(ai_extractor.ai_skill_counter),
        'top_ai_skills': ai_extractor.ai_skill_counter.most_common(20),
        'extraction_method': 'AI (OpenAI GPT)'
    })

@app.route('/api/comparison')
def api_comparison():
    """API endpoint to compare pattern matching vs AI extraction results."""
    pattern_skills = set(skill_counter.keys())
    ai_skills = set(ai_extractor.ai_skill_counter.keys())
    
    return jsonify({
        'pattern_matching': {
            'total_skills': len(pattern_skills),
            'top_skills': skill_counter.most_common(10),
            'total_documents': len(processed_documents)
        },
        'ai_extraction': {
            'total_skills': len(ai_skills),
            'top_skills': ai_extractor.ai_skill_counter.most_common(10),
            'total_documents': len(ai_extractor.ai_processed_documents),
            'service_type': ai_extractor.service_type,
            'model_name': ai_extractor.model_name,
            'debug_info': {
                'counter_length': len(ai_extractor.ai_skill_counter),
                'has_ai_client': ai_extractor.client is not None,
                'blob_name': ai_extractor.ai_stats_blob_name
            }
        },
        'comparison': {
            'common_skills': len(pattern_skills.intersection(ai_skills)),
            'pattern_only': len(pattern_skills - ai_skills),
            'ai_only': len(ai_skills - pattern_skills),
            'overlap_percentage': round(len(pattern_skills.intersection(ai_skills)) / max(len(pattern_skills.union(ai_skills)), 1) * 100, 2)
        }
    })

@app.route('/api/health')
def health_check():
    """Health check endpoint with stats information."""
    stats_loaded = len(skill_counter) > 0 or len(processed_documents) > 0
    
    return jsonify({
        'status': 'healthy',
        'stats_loaded': stats_loaded,
        'total_documents': len(processed_documents),
        'total_skills': len(skill_counter),
        'azure_blob_available': get_blob_service_client() is not None,
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/reload-stats', methods=['POST'])
def reload_stats():
    """Manually reload stats from Azure Blob Storage."""
    try:
        success = load_stats_from_blob()
        if success:
            return jsonify({
                'success': True,
                'message': 'Stats reloaded successfully',
                'total_documents': len(processed_documents),
                'total_skills': len(skill_counter)
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to reload stats from blob storage'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error reloading stats: {str(e)}'
        })

@app.route('/skills')
def skills_page():
    """Page showing all skill statistics."""
    all_skills = skill_counter.most_common()
    all_ai_skills = ai_extractor.ai_skill_counter.most_common()
    
    # Create combined skill data with both pattern and AI counts
    combined_skills = {}
    
    # Add pattern matching skills
    for skill, count in all_skills:
        combined_skills[skill] = {'pattern_count': count, 'ai_count': 0}
    
    # Add AI skills
    for skill, count in all_ai_skills:
        if skill in combined_skills:
            combined_skills[skill]['ai_count'] = count
        else:
            combined_skills[skill] = {'pattern_count': 0, 'ai_count': count}
    
    # Sort by total occurrences (pattern + AI)
    sorted_skills = sorted(combined_skills.items(), 
                          key=lambda x: x[1]['pattern_count'] + x[1]['ai_count'], 
                          reverse=True)
    
    return render_template('skills.html', 
                         skills=all_skills, 
                         combined_skills=sorted_skills,
                         page_name='skills')

@app.route('/ai-skills')
def ai_skills_page():
    """Page showing AI-extracted skill statistics."""
    all_ai_skills = ai_extractor.ai_skill_counter.most_common()
    return render_template('ai_skills.html', ai_skills=all_ai_skills, page_name='ai-skills')

@app.route('/skill-details')
def skill_details():
    """Page showing skills with document references."""
    # Create a comprehensive view of skills with their documents
    skills_with_docs = []
    
    for skill, count in skill_counter.most_common():
        documents = skill_documents.get(skill, [])
        skills_with_docs.append({
            'skill': skill,
            'count': count,
            'documents': documents
        })
    
    return render_template('skill_details.html', skills_data=skills_with_docs, page_name='skill-details')

@app.route('/comparison')
def comparison_page():
    """Page comparing pattern matching vs AI extraction results."""
    # Get data for comparison
    pattern_skills = skill_counter.most_common(20)
    ai_skills = ai_extractor.ai_skill_counter.most_common(20)
    
    # Get chart data for both methods
    pattern_chart_data = get_monthly_chart_data()
    ai_chart_data = ai_extractor.get_ai_monthly_chart_data()
    
    # Calculate comparison statistics
    pattern_skill_names = set([skill for skill, count in pattern_skills])
    ai_skill_names = set([skill for skill, count in ai_skills])
    
    # Get all skills from both methods (not just top 20)
    all_pattern_skills = set([skill for skill, count in skill_counter.most_common()])
    all_ai_skills = set([skill for skill, count in ai_extractor.ai_skill_counter.most_common()])
    
    # Calculate overlap and unique skills
    common_skills = all_pattern_skills & all_ai_skills
    unique_pattern_skills = list(all_pattern_skills - all_ai_skills)
    unique_ai_skills = list(all_ai_skills - all_pattern_skills)
    total_unique = len(all_pattern_skills | all_ai_skills)
    overlap_count = len(common_skills)
    
    return render_template('comparison.html', 
                         pattern_skills=pattern_skills,
                         ai_skills=ai_skills,
                         pattern_chart_data=pattern_chart_data,
                         ai_chart_data=ai_chart_data,
                         overlap_count=overlap_count,
                         total_unique=total_unique,
                         unique_pattern_skills=unique_pattern_skills,
                         unique_ai_skills=unique_ai_skills,
                         page_name='comparison')

@app.route('/about')
def about_page():
    """About page with application information."""
    # Calculate dynamic statistics
    total_skills_in_db = len(tech_skills)
    unique_skills_found = len(skill_counter)
    total_documents = len(processed_documents)
    total_skill_occurrences = sum(skill_counter.values())
    
    # Calculate categories (this is an approximation based on the skills.py structure)
    categories = [
        "Programming Languages", "Web Technologies", "Databases", "Cloud Platforms",
        "DevOps & Tools", "Mobile Development", "Data Science & AI", "Testing",
        "Operating Systems", "Version Control", "IDEs & Editors", "Methodologies",
        "Other Technologies"
    ]
    total_categories = len(categories)
    
    return render_template('about.html', 
                         total_skills_in_db=total_skills_in_db,
                         unique_skills_found=unique_skills_found,
                         total_documents=total_documents,
                         total_skill_occurrences=total_skill_occurrences,
                         total_categories=total_categories,
                         page_name='about')

@app.route('/debug/sync-ai', methods=['GET'])
def debug_sync_ai():
    """Debug route to force sync AI extractor with processed documents."""
    try:
        print("Debug: Force syncing AI extractor...")
        success = force_sync_ai_extractor_with_processed_documents()
        
        stats = {
            'sync_success': success,
            'ai_skills_count': len(ai_extractor.ai_skill_counter),
            'ai_documents_count': len(ai_extractor.ai_processed_documents),
            'processed_docs_with_ai': len([d for d in processed_documents.values() if d.get('ai_skills_found')]),
            'top_ai_skills': ai_extractor.ai_skill_counter.most_common(10) if ai_extractor.ai_skill_counter else []
        }
        
        return jsonify(stats)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Azure Blob Storage helper functions
# Stats persistence functions
def save_stats_to_blob():
    """Save all statistics to Azure Blob Storage as JSON files."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            return False
        
        stats_data = {
            'skill_counter': dict(skill_counter),
            'skill_documents': dict(skill_documents),
            'processed_documents': dict(processed_documents),
            'monthly_skill_data': {k: dict(v) for k, v in monthly_skill_data.items()},
            'last_updated': datetime.now().isoformat()
        }
        
        # Upload stats as JSON
        blob_client = blob_service_client.get_blob_client(
            container=AZURE_STORAGE_CONTAINER_NAME,
            blob='stats/app_statistics.json'
        )
        
        blob_client.upload_blob(
            json.dumps(stats_data, indent=2).encode('utf-8'),
            overwrite=True
        )
        return True
    except Exception as e:
        print(f"Error saving stats to blob: {e}")
        return False

def load_stats_from_blob():
    """Load statistics from Azure Blob Storage JSON files."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            return False
        
        blob_client = blob_service_client.get_blob_client(
            container=AZURE_STORAGE_CONTAINER_NAME,
            blob='stats/app_statistics.json'
        )
        
        stats_json = blob_client.download_blob().readall().decode('utf-8')
        stats_data = json.loads(stats_json)
        
        # Restore global variables
        global skill_counter, skill_documents, processed_documents, monthly_skill_data
        
        skill_counter = Counter(stats_data.get('skill_counter', {}))
        skill_documents = defaultdict(list, stats_data.get('skill_documents', {}))
        processed_documents = stats_data.get('processed_documents', {})
        monthly_skill_data = defaultdict(lambda: defaultdict(int))
        
        # Restore monthly data structure
        for skill, months in stats_data.get('monthly_skill_data', {}).items():
            for month, count in months.items():
                monthly_skill_data[skill][month] = count
        
        print(f"Stats loaded successfully. Last updated: {stats_data.get('last_updated', 'Unknown')}")
        return True
    except Exception as e:
        print(f"Error loading stats from blob: {e}")
        return False

# Load stats on application startup
# Try to load stats when the application starts
try:
    load_stats_from_blob()
    print("Application stats loaded successfully on startup")
except Exception as e:
    print(f"Failed to load stats on startup: {e}")
    print("Starting with empty stats")

def upload_file_to_blob(file_content, filename):
    """Upload file to Azure Blob Storage."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            return False, "Azure Blob Storage not configured"
        
        blob_client = blob_service_client.get_blob_client(
            container=AZURE_STORAGE_CONTAINER_NAME,
            blob=filename
        )
        
        blob_client.upload_blob(file_content, overwrite=True)
        return True, "File uploaded successfully"
    except Exception as e:
        print(f"Error uploading to blob storage: {e}")
        return False, f"Upload failed: {str(e)}"

def download_file_from_blob(filename):
    """Download file from Azure Blob Storage."""
    try:
        blob_service_client = get_blob_service_client()
        if not blob_service_client:
            return None
        
        blob_client = blob_service_client.get_blob_client(
            container=AZURE_STORAGE_CONTAINER_NAME,
            blob=filename
        )
        
        return blob_client.download_blob().readall()
    except Exception as e:
        print(f"Error downloading from blob storage: {e}")
        return None

def save_file_safely(file, filename):
    """Save file to Azure Blob Storage or local storage (fallback)."""
    file_content = file.read()
    
    # Try Azure Blob Storage first
    blob_service_client = get_blob_service_client()
    if blob_service_client:
        success, message = upload_file_to_blob(file_content, filename)
        if success:
            return file_content, True  # Return content and success flag
    
    # Fallback to local storage
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    with open(filepath, 'wb') as f:
        f.write(file_content)
    
    return file_content, False  # Return content and blob flag (False = local)

def get_file_content(filename, is_blob=True):
    """Get file content from Azure Blob Storage or local storage."""
    if is_blob:
        content = download_file_from_blob(filename)
        if content:
            return content
    
    # Fallback to local storage
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        with open(filepath, 'rb') as f:
            return f.read()
    except Exception as e:
        print(f"Error reading local file: {e}")
        return None

@app.route('/api/reload-ai-stats', methods=['POST'])
def reload_ai_stats():
    """Manually reload AI stats from blob storage for debugging."""
    try:
        # First try to sync from processed documents
        sync_success = sync_ai_extractor_with_processed_documents()
        
        # Then try to load from dedicated AI blob if needed
        blob_success = False
        if len(ai_extractor.ai_skill_counter) == 0:
            blob_success = load_ai_stats_from_blob()
        
        total_skills = len(ai_extractor.ai_skill_counter)
        total_docs = len(ai_extractor.ai_processed_documents)
        
        if sync_success or blob_success or total_skills > 0:
            return jsonify({
                'success': True,
                'message': f'AI stats reloaded successfully (sync: {sync_success}, blob: {blob_success})',
                'ai_stats': {
                    'total_documents': total_docs,
                    'total_skills': total_skills,
                    'top_skills': ai_extractor.ai_skill_counter.most_common(10),
                    'sync_from_docs': sync_success,
                    'loaded_from_blob': blob_success
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Failed to reload AI stats from both sync and blob',
                'ai_stats': {
                    'total_documents': total_docs,
                    'total_skills': total_skills,
                    'sync_from_docs': sync_success,
                    'loaded_from_blob': blob_success
                }
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error reloading AI stats: {str(e)}'
        })

@app.route('/api/monthly-analysis')
def get_monthly_analysis():
    """Get the latest monthly analysis report."""
    try:
        latest_report = monthly_analyzer.get_latest_report()
        
        if latest_report:
            return jsonify({
                'success': True,
                'report': latest_report
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No monthly reports available'
            })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error retrieving monthly analysis: {str(e)}'
        })

@app.route('/api/generate-monthly-analysis', methods=['POST'])
def generate_monthly_analysis():
    """Generate a new monthly analysis report."""
    try:
        # Safely get JSON data, default to empty dict if not provided
        data = request.get_json(silent=True) or {}
        target_month = data.get('month') if data else None
        
        # Generate the report
        report = monthly_analyzer.generate_monthly_report(target_month)
        
        return jsonify({
            'success': True,
            'message': 'Monthly analysis generated successfully',
            'report': {
                'analysis_month': report.analysis_month,
                'total_documents': report.total_documents,
                'total_resumes': report.total_resumes,
                'total_job_descriptions': report.total_job_descriptions,
                'top_technical_skills': report.top_technical_skills[:10],
                'top_soft_skills': report.top_soft_skills[:10],
                'emerging_skills': report.emerging_skills[:5],
                'recommendations': report.recommendations,
                'generated_at': report.generated_at
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generating monthly analysis: {str(e)}'
        })

@app.route('/api/monthly-analysis/historical')
def get_historical_monthly_reports():
    """Get index of all historical monthly reports."""
    try:
        historical_data = monthly_analyzer.get_historical_reports()
        
        return jsonify({
            'success': True,
            'historical_data': historical_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error retrieving historical reports: {str(e)}'
        })

@app.route('/api/monthly-analysis/<month>')
def get_monthly_report_by_month(month):
    """Get a specific monthly report by month (YYYY-MM format)."""
    try:
        # Validate month format
        if not month or len(month) != 7 or month[4] != '-':
            return jsonify({
                'success': False,
                'message': 'Invalid month format. Use YYYY-MM format (e.g., 2025-09)'
            }), 400
        
        report = monthly_analyzer.get_report_by_month(month)
        
        if report:
            return jsonify({
                'success': True,
                'report': report
            })
        else:
            return jsonify({
                'success': False,
                'message': f'No report found for {month}'
            }), 404
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error retrieving report for {month}: {str(e)}'
        })

@app.route('/api/monthly-analysis/compare', methods=['POST'])
def compare_monthly_reports():
    """Compare multiple monthly reports."""
    try:
        data = request.get_json()
        months = data.get('months', [])
        
        if not months or len(months) < 2:
            return jsonify({
                'success': False,
                'message': 'Please provide at least 2 months for comparison'
            }), 400
        
        comparison = monthly_analyzer.get_comparative_analysis(months)
        
        return jsonify({
            'success': True,
            'comparison': comparison
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error generating comparison: {str(e)}'
        })

@app.route('/monthly-dashboard')
def monthly_dashboard():
    """Display monthly analysis dashboard."""
    try:
        latest_report = monthly_analyzer.get_latest_report()
        historical_data = monthly_analyzer.get_historical_reports()
        
        return render_template('monthly_dashboard.html', 
                             report=latest_report, 
                             historical_data=historical_data)
    except Exception as e:
        return render_template('monthly_dashboard.html', error=str(e))

@app.route('/debug/monthly-reports')
def debug_monthly_reports():
    """Debug endpoint to view monthly reports status."""
    try:
        historical_data = monthly_analyzer.get_historical_reports()
        latest_report = monthly_analyzer.get_latest_report()
        
        debug_info = {
            'historical_summary': {
                'total_reports': historical_data.get('total_reports', 0),
                'months_available': historical_data.get('months_available', [])
            },
            'latest_report_month': latest_report.get('analysis_month', 'None'),
            'latest_report_documents': latest_report.get('total_documents', 0),
            'azure_storage_container': monthly_analyzer.reports_container,
            'blob_prefix': monthly_analyzer.reports_blob_prefix,
            'historical_index_blob': monthly_analyzer.historical_index_blob
        }
        
        return jsonify(debug_info)
    except Exception as e:
        return jsonify({
            'error': str(e),
            'debug_info': 'Failed to load monthly reports debug information'
        })

if __name__ == '__main__':
    app.run(debug=True, port=5001)
